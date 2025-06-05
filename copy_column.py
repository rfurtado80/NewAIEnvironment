import os
from openpyxl import load_workbook


def main():
    filepath = input("Enter the path to the Excel file: ").strip()
    if not os.path.isfile(filepath):
        print(f"File not found: {filepath}")
        return

    wb = load_workbook(filepath)
    ws = wb.active

    # Copy values from column B to column C
    for row in ws.iter_rows(min_col=2, max_col=2):
        cell_b = row[0]
        ws.cell(row=cell_b.row, column=3, value=cell_b.value)

    dirpath = os.path.dirname(filepath)
    filename = os.path.basename(filepath)
    name, ext = os.path.splitext(filename)
    output_path = os.path.join(dirpath, f"{name}_modified{ext}")
    wb.save(output_path)
    print(f"Saved modified file to {output_path}")


if __name__ == "__main__":
    main()
